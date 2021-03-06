#!/usr/bin/python3
#################################################################################
#  Script     : bank_recon
#  Author     : Farhad Ali
#  Description: This script highlights matching transcations in an excel sheet
#               It requires two sheets to do so:-
#               (i).    usersheet - holds user records
#               (ii).   banksheet - holds records from bank reconciliation report
#################################################################################
# -*- coding: UTF-8 -*
##############################################################################
# imports required for the script to work
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
import sys
import time
import os
import glob
import cgi
import html
import cgitb; cgitb.enable()     # for troubleshooting


if sys.version[0] == '2':
    reload(sys)
    sys.setdefaultencoding("utf-8")

###############################################################################

os.chdir(".")       # bound path to the current directory

#satisfy openpyxl requirements for highlighting cells
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=12)
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = PatternFill(fill_type='lightUp',
                 start_color='fff000',
                end_color='6efdfd')

print("\n")
print("files avilable in this folder:")
print("\n")

#enum files in current directory
files = os.listdir('/var/www/cgi-bin')
i = 1
for f in glob.glob("*.xlsx"):
    print("(" + str(i) + "). "  + str(f))
    i+= 1

print("\n")

print("Content-Type: text/html") # HTTP header to say HTML is following
print()                          # blank line, end of headers

form = cgi.FieldStorage()
wb_name  = html.escape(form["wb_name"].value);
u_sheet  = html.escape(form["u_sheet"].value);
b_sheet  = html.escape(form["b_sheet"].value);

#print("Adicione o nome da planilha em a extensao:")
#wb_name = input()
print("\n")
print(wb_name)
print(b_sheet)
print(u_sheet)
time.sleep(1)
try:
    workBook = openpyxl.load_workbook(wb_name + str(".xlsx"))
except IOError:
    print("could not find the book. exiting...")
    sys.exit()
print("I found the following sheets in your book:")
print("\n")
j = 1;
for sheets in workBook.sheetnames:
    print("(" + str(j) + "). "  + str(sheets))
    j += 1
print("\n")
print("enter the sheet name with bank data:")
b_sheet = input()
print("\n")
try:
    bankSheet = workBook.get_sheet_by_name(b_sheet)
except KeyError:
    print("no such sheet in your file: " + wb_name + " xlsx.")
    print("exitting....")
    sys.exit()

print("SUCESS: bank data found at sheet: " + b_sheet)
print("\n")
print("enter the sheet name with your data:")
u_sheet = input()
try:
    userSheet = workBook.get_sheet_by_name(u_sheet)
    print("\n")
    print("SUCESS: user data found at sheet: " + u_sheet)
    print("\n")
except KeyError:
    print("no such sheet found in your file: " + wb_name + " xlsx.")
    sys.exit()

# def get_dates(sheetName):
#     dates = []
#     for row in range(2, sheetName.max_row + 1):
#         cellObj = sheetName["A" + str(row)]
#         eachDate = cellObj.value
#         dates.append(eachDate)
#     return dates
#
# #loop through all records in Column B of the excel file and convert them
# #into an array. return that array
# def get_chqs(sheetName):
#     chqs = []
#     for row in range(2, sheetName.max_row + 1):
#         cellObj = sheetName["B" + str(row)]
#         eachChq = cellObj.value
#         chqs.append(eachChq)
#     return chqs
#
# def get_amount(sheetName):
#     amount = []
#     for row in range(2, sheetName.max_row + 1):
#         cellObj = sheetName["C" + str(row)]
#         eachAmu = cellObj.value
#         amount.append(eachAmu)
#     return amount
#
# dates = get_dates(bankSheet)
# cheques = get_chqs(bankSheet)
# amounts = get_amount(bankSheet)


def get_trios(sheetName):
    trios = []

    for row in range(2, sheetName.max_row + 1):
        _date = sheetName["A" + str(row)]
        desc = sheetName["B" + str(row)]
        val = sheetName["C" + str(row)]
        trios.append((_date.value, val.value))

    return trios


trios = get_trios(sheetName=bankSheet)


# print("Number of total dates found: " + str(len(dates)))
# print("Number of total cheques found: " + str(len(cheques)))
# print("Number of total amounts found: " + str(len(amounts)))
print("\n")
print("do you want to see the cheques & amounts I found in? " + b_sheet + "(Y/N)")

while True:
    see_chq = input()
    if see_chq == "Y":
        break
    elif see_chq == "N":
        print("\n")
        print("not displaying cheques & amounts..")
        print("going onward....")
        time.sleep(2)
        break
    else:
        print("Y or N")


print("processing your data....")
time.sleep(2)
print("finding matches...")
time.sleep(2)
count = 0   #keep track of matches found
for row in range(2, userSheet.max_row + 1):
    DtcellObject = userSheet["A" + str(row)]
    ChcellObject = userSheet["B" + str(row)]        #get cell Object for every record found on column B of excel sheet
    AmcellObject = userSheet["C" + str(row)]        #same as above for every record on column C of the excel file

    trio = (DtcellObject.value, AmcellObject.value)

    if trio in trios:
        DtcellObject.style = highlight
        ChcellObject.style = highlight
        AmcellObject.style = highlight
        count += 1

    # if ChcellObject.value in cheques:               #check for matches in the "cheque" column
    #     if AmcellObject.value in amounts:           #check  for matches in the "amount" column
    #         AmcellObject.style = highlight          # highlight them all :)
    #         ChcellObject.style = highlight          # --do--
    #     count += 1

print(str(count) + " matches found")
print("\n")
print("hold on...")
time.sleep(1)
print("highlighting in process...")
time.sleep(2)
print("SUCCESS:" + str(count) + " matches highlighted")
time.sleep(1)
print("creating new file in your folder....")
time.sleep(1)
workBook.save("conciliacao.xlsx")             # create new file with all the matched instance highlighted automatically
print("ready.xlsx created")
time.sleep(2)
print("this script was created by Farhad Ali. Email: alifarhad557@gmail.com")
time.sleep(2)
print("Exiting...")
